import tkinter as tk
from tkinter import messagebox
from PIL import Image, ImageTk
import openpyxl
from openpyxl import Workbook
from tkcalendar import DateEntry

def submit_form():
    global student_name_entry, father_name_entry, mother_name_entry, age_entry, contact_number_entry
    global category_combobox, gender_var, branch_combobox, dob_entry, email_entry

    # Retrieve form data
    student_name = student_name_entry.get()
    father_name = father_name_entry.get()
    mother_name = mother_name_entry.get()
    age = age_entry.get()
    contact_number = contact_number_entry.get()
    category = category_combobox.get()
    gender = gender_var.get()
    branch = branch_combobox.get()
    dob = dob_entry.get()
    
    email = email_entry.get()

    # Check if required fields are empty
    if not student_name or not age or not contact_number:
        messagebox.showerror("Error", "Please fill in all the required fields (Name, Age and Mobile Number).")
        return

    # Open the Excel file
    try:
        workbook = openpyxl.load_workbook("admission_data.xlsx")
    except FileNotFoundError:
        workbook = Workbook()

    # Select the active sheet
    sheet = workbook.active

    # If the sheet is empty, add the column headers
    if sheet.max_row == 1 and sheet.max_column == 1:
        headers = ["Student Name", "Father Name", "Mother Name", "Age", "Contact Number", "Category", "Gender", "Branch", "Date of Birth", "Email ID"]
        for col_num, header in enumerate(headers, 1):
            sheet.cell(row=1, column=col_num).value = header

    # Find the next available row
    next_row = sheet.max_row + 1

    # Write the form data to the corresponding columns
    sheet.cell(row=next_row, column=1).value = student_name
    sheet.cell(row=next_row, column=2).value = father_name
    sheet.cell(row=next_row, column=3).value = mother_name
    sheet.cell(row=next_row, column=4).value = age
    sheet.cell(row=next_row, column=5).value = contact_number
    sheet.cell(row=next_row, column=6).value = category
    sheet.cell(row=next_row, column=7).value = gender
    sheet.cell(row=next_row, column=8).value = branch
    sheet.cell(row=next_row, column=9).value = dob
    sheet.cell(row=next_row, column=10).value = email

    # Save the changes to the Excel file
    workbook.save("admission_data.xlsx")

    # Display a message box with the form data
    messagebox.showinfo("Form Submission", "Form data has been saved successfully.")

def open_admission_form():
    
    global student_name_entry, father_name_entry, mother_name_entry, age_entry, contact_number_entry
    global category_combobox, gender_var, branch_combobox, dob_entry, email_entry

    admission_window = tk.Toplevel()
    admission_window.title("Admission Form")
    admission_window.geometry("400x600")

    student_name_label = tk.Label(admission_window, text="Student Name: *")
    student_name_label.pack()
    student_name_entry = tk.Entry(admission_window)
    student_name_entry.pack()
    student_name_entry.config(validate="key")
    student_name_entry.config(validatecommand=(student_name_entry.register(lambda text: text.isalpha() or text == ""), "%P"))

    father_name_label = tk.Label(admission_window, text="Father Name:")
    father_name_label.pack()
    father_name_entry = tk.Entry(admission_window)
    father_name_entry.pack()
    father_name_entry.config(validate="key")
    father_name_entry.config(validatecommand=(father_name_entry.register(lambda text: text.isalpha() or text == ""), "%P"))

    mother_name_label = tk.Label(admission_window, text="Mother Name:")
    mother_name_label.pack()
    mother_name_entry = tk.Entry(admission_window)
    mother_name_entry.pack()
    mother_name_entry.config(validate="key")
    mother_name_entry.config(validatecommand=(mother_name_entry.register(lambda text: text.isalpha() or text == ""), "%P"))

    age_label = tk.Label(admission_window, text="Age: *")
    age_label.pack()
    age_entry = tk.Entry(admission_window)
    age_entry.pack()
    age_entry.config(validate="key")
    age_entry.config(validatecommand=(age_entry.register(lambda text: text.isdigit() and len(text) <= 2), "%P"))

    contact_number_label = tk.Label(admission_window, text="Contact Number: *")
    contact_number_label.pack()
    contact_number_entry = tk.Entry(admission_window)
    contact_number_entry.pack()
    contact_number_entry.config(validate="key")
    contact_number_entry.config(validatecommand=(contact_number_entry.register(lambda text: text.isdigit() and len(text) <= 10), "%P"))

    category_label = tk.Label(admission_window, text="Category:")
    category_label.pack()
    category_combobox = tk.StringVar()
    category_combobox.set("General")
    category_dropdown = tk.OptionMenu(admission_window, category_combobox, "General", "OBC", "SC", "ST")
    category_dropdown.pack()

    gender_label = tk.Label(admission_window, text="Gender:")
    gender_label.pack()
    gender_var = tk.StringVar()
    gender_var.set("Male")
    gender_radio1 = tk.Radiobutton(admission_window, text="Male", variable=gender_var, value="Male")
    gender_radio1.pack()
    gender_radio2 = tk.Radiobutton(admission_window, text="Female", variable=gender_var, value="Female")
    gender_radio2.pack()

    branch_label = tk.Label(admission_window, text="Branch:")
    branch_label.pack()
    branch_combobox = tk.StringVar(admission_window)
    branch_combobox.set("ME")
    branch_dropdown = tk.OptionMenu(admission_window, branch_combobox, "ME", "ECE", "EE", "CSE", "CE")
    branch_dropdown.pack()

    dob_label = tk.Label(admission_window, text="Date of Birth:")
    dob_label.pack()
    dob_entry = DateEntry(admission_window, date_pattern='dd/mm/yyyy')
    dob_entry.pack()


    email_label = tk.Label(admission_window, text="Email:")
    email_label.pack()
    email_entry = tk.Entry(admission_window)
    email_entry.pack()

    submit_button = tk.Button(admission_window, text="Submit", command=submit_form)
    submit_button.pack()

    admission_window.mainloop()
